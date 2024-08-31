# pandasを使ったエクセルシートの操作
import pandas as pd
from openpyxl import load_workbook

# Excelファイルのパスとシート名を指定
# file_path = './data/abc.xlsx'
# sheet_name = 'sheet1'

# ---Excel操作の基本 begin-------------------------------
def case_pandas(file_path, sheet_name):
    try:
        # Excelシートを読み込む（ヘッダーなしとして読み込む）
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)

        print("Original DataFrame columns:", df.columns)

        # 列名を手動で割り当てる
        df.columns = ['A', 'B', 'C']

        print("Updated DataFrame columns:", df.columns)

        df.at[0, 'A'] = 1
        df.at[0, 'B'] = 2

        # A1 + B1 の結果を C1 に格納
        df.at[0, 'C'] = df.at[0, 'A'] + df.at[0, 'B']

        # Excelファイルに保存
        df.to_excel(file_path, sheet_name=sheet_name, index=False, header=False)
        print("Excel file updated successfully.")

        # 更新後のデータを表示
        print("Updated data:")
        print(df)
    except Exception as e:
        print(f"An error occurred: {str(e)}")

def case_openpyxl(file_path, sheet_name):
    # Openpyxlで既存のExcelファイルを読み込み
    wb = load_workbook(file_path)
    ws = wb[sheet_name]

    # セルC1に計算式を設定
    ws['C1'] = '=A1+B1'

    # Excelファイルに保存
    wb.save(file_path)
# ---Excel操作の基本 end -------------------------------

def chat_completions_on_excel():

    messages = []
    system_content = "あなたは有能なソフトウェア開発者のアシスタントです。"
    user_content = "プロのソフトウェア開発者向けに、OpenAiのAPIの概要を説明しなさい。"

    prompt = ""
    completions = ""



def main():
    file_path = './data/abc.xlsx'
    sheet_name = 'Sheet1'
    case_pandas(file_path, sheet_name)

    file_path = './data/abc2.xlsx'
    sheet_name = 'Sheet1'
    case_openpyxl(file_path, sheet_name)

if __name__ == '__main__':
    main()
